from flask import Flask, render_template, jsonify, request
import json
import time
import threading
from datetime import datetime, timedelta
import os
import random
import pandas as pd
import glob
from openpyxl import load_workbook
from config import config

# 获取环境配置
env = os.environ.get('FLASK_ENV', 'development')
app = Flask(__name__)
app.config.from_object(config[env])

# 订单数据存储
orders_db = []
order_counter = 1
# 记录前端操作的时间戳，防止被Excel数据覆盖
frontend_operations = {}  # {order_id: last_operation_time}
# Excel文件监控变量
excel_file_modified_time = None  # 记录Excel文件最后修改时间
is_excel_updating = False  # 标记Excel是否正在被外部程序更新

# 订单状态常量
PENDING = 2          # 备货中
COMPLETED = 5        # 已完成

# Excel文件路径配置
# 获取项目根目录的data文件夹路径
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FOLDER = os.path.join(PROJECT_ROOT, "data")  # 从项目data文件夹读取Excel文件
EXCEL_PATTERN = "*前端咖啡订单*.xlsx"  # 只匹配前端专用咖啡订单Excel文件

# Excel文件操作锁，防止并发写入冲突
excel_write_lock = threading.Lock()

def ensure_orders_folder():
    """确保桌面路径存在"""
    if not os.path.exists(EXCEL_FOLDER):
        print(f"❌ 桌面路径不存在: {EXCEL_FOLDER}")
        return False
    print(f"✅ 桌面路径正常: {EXCEL_FOLDER}")
    return True

def ensure_excel_files_writable():
    """确保Excel文件可写"""
    try:
        # 查找所有咖啡订单Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if not excel_files:
            print("📁 未找到咖啡订单Excel文件，跳过权限检查")
            return
        
        print("🔧 检查咖啡订单Excel文件写入权限...")
        
        for file_path in excel_files:
            try:
                # 检查文件是否可写
                if not os.access(file_path, os.W_OK):
                    print(f"⚠️  文件不可写: {os.path.basename(file_path)}")
                    
                    # 尝试修复文件权限
                    import stat
                    current_attrs = os.stat(file_path).st_mode
                    
                    # 添加写入权限
                    new_attrs = current_attrs | stat.S_IWRITE
                    os.chmod(file_path, new_attrs)
                    
                    print(f"✅ 已修复文件权限: {os.path.basename(file_path)}")
                else:
                    print(f"✅ 文件可写: {os.path.basename(file_path)}")
                    
            except Exception as e:
                print(f"❌ 处理文件权限时出错 {os.path.basename(file_path)}: {e}")
        
        print("🎉 咖啡订单Excel文件权限检查完成")
        
    except Exception as e:
        print(f"❌ 检查Excel文件权限时出错: {e}")

def map_order_status(status_text):
    """映射订单状态文本到数字状态"""
    status_map = {
        '备货中': '2',
        '已完成': '5',
        '2': '2',
        '5': '5'
    }
    return status_map.get(str(status_text), str(status_text))

def read_excel_orders():
    """从Excel文件读取订单数据"""
    global orders_db, excel_file_modified_time, is_excel_updating
    
    try:
        # 检查是否有新数据标志
        flag_file_path = os.path.join(EXCEL_FOLDER, "new_data_ready.flag")
        if os.path.exists(flag_file_path):
            print("🚩 检测到新数据标志，开始处理...")
            # 创建锁文件，通知后端正在处理
            lock_file_path = os.path.join(EXCEL_FOLDER, "data.lock")
            try:
                with open(lock_file_path, 'w') as f:
                    json.dump({
                        'pid': os.getpid(),
                        'timestamp': time.time(),
                        'process': 'frontend_processing'
                    }, f)
                print("🔒 已创建前端处理锁")
            except Exception as e:
                print(f"⚠️ 创建锁文件失败: {e}")
        
        # 检查文件锁
        lock_file_path = os.path.join(EXCEL_FOLDER, "data.lock")
        if os.path.exists(lock_file_path):
            try:
                with open(lock_file_path, 'r') as f:
                    lock_info = json.load(f)
                    lock_time = lock_info.get('timestamp', 0)
                    # 如果锁文件超过30秒，可能是僵尸锁，可以忽略
                    if time.time() - lock_time < 30:
                        print("⚠️ 检测到文件锁，等待数据更新完成...")
                        return  # 跳过本次读取
            except:
                pass  # 如果读取锁文件失败，继续正常流程
        
        # 确保桌面路径存在
        if not ensure_orders_folder():
            print("❌ 无法访问桌面路径，使用空订单列表")
            orders_db = []
            return
        
        # 确保Excel文件可写
        ensure_excel_files_writable()
        
        # 查找所有咖啡订单Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if not excel_files:
            print("未找到咖啡订单Excel文件，使用空订单列表")
            orders_db = []
            return
        
        # 读取最新的咖啡订单Excel文件
        latest_file = max(excel_files, key=os.path.getctime)
        
        # 检查文件是否存在
        if not os.path.exists(latest_file):
            print(f"⚠️ 文件不存在，可能被删除: {latest_file}")
            orders_db = []
            return
        
        # 检查Excel文件是否被外部程序修改
        current_modified_time = os.path.getmtime(latest_file)
        
        if excel_file_modified_time is not None and current_modified_time > excel_file_modified_time:
            # 咖啡订单Excel文件被外部程序修改了
            print(f"🔄 检测到咖啡订单Excel文件被外部程序修改: {latest_file}")
            is_excel_updating = True
            
            # 等待一段时间，确保外部程序完成写入
            time.sleep(3)
            
            # 再次检查修改时间，确保写入完成
            final_modified_time = os.path.getmtime(latest_file)
            if final_modified_time == current_modified_time:
                print("✅ 咖啡订单Excel文件写入完成，开始读取新数据")
                # 保持冲突检测状态5秒，防止用户操作冲突
                time.sleep(5)
                is_excel_updating = False
            else:
                print("⏳ 咖啡订单Excel文件仍在被修改，等待完成...")
                time.sleep(3)
                is_excel_updating = False
        
        # 更新文件修改时间
        excel_file_modified_time = current_modified_time
        
        print(f"读取咖啡订单Excel文件: {latest_file}")
        
        # 读取Excel数据
        try:
            df = pd.read_excel(latest_file, engine='openpyxl')
            print(f"成功读取咖啡订单Excel文件，数据行数: {len(df)}")
        except FileNotFoundError:
            print(f"⚠️ 文件被删除，等待重新生成: {latest_file}")
            # 等待一段时间，让后端重新生成文件
            time.sleep(2)
            return
        except Exception as e:
            print(f"❌ 读取Excel文件失败: {e}")
            return
        
        # 保存现有订单数据用于比较
        old_orders_db = orders_db.copy()
        # 清空现有订单数据
        orders_db = []
        
        # 处理每一行数据
        valid_order_id = 1
        for index, row in df.iterrows():
            try:
                # 检查是否为空行
                order_number = row.get('订单编号')
                if pd.isna(order_number):
                    continue
                
                # 使用订单编号查找现有订单，而不是依赖ID
                existing_order_by_number = next((o for o in old_orders_db if o['number'] == str(order_number)), None)
                
                # 检查是否有前端操作记录（基于订单编号）
                has_frontend_operation = str(order_number) in [str(op_id) for op_id in frontend_operations.keys()]
                frontend_operation = None
                if has_frontend_operation:
                    # 找到对应的操作记录
                    for op_id, op_data in frontend_operations.items():
                        if str(op_id) == str(order_number):
                            frontend_operation = op_data
                            break
                
                # 获取订单状态
                status_text = str(row.get('订单状态', '备货中'))
                
                # 跳过已取消的订单
                if status_text == '已取消':
                    continue
                
                # 根据前端Excel格式映射数据
                status_code = str(row.get('状态', '2'))
                status_name = str(row.get('状态名称', '备货中'))
                
                # 如果状态名称为空，根据状态代码生成
                if not status_name or status_name == 'nan':
                    if status_code == '5':
                        status_name = '已完成'
                    else:
                        status_name = '备货中'
                
                # 如果是现有订单，使用原有ID；否则分配新ID
                if existing_order_by_number:
                    order_id = existing_order_by_number['id']
                else:
                    # 为新订单分配ID
                    if old_orders_db:
                        order_id = max(o['id'] for o in old_orders_db) + 1
                    else:
                        order_id = 1
                
                order = {
                    'id': order_id,
                    'number': str(order_number),
                    'status': status_name,  # 使用状态名称而不是代码
                    'userName': str(row.get('姓名', '未知')),
                    'phone': str(row.get('手机号码', '未知')),
                    'address': str(row.get('部门', '未知')),
                    'amount': 0.0,  # 前端Excel中没有金额字段
                    'orderTime': str(row.get('支付时间', datetime.now().isoformat())),
                    'remark': str(row.get('订单备注', 'nan')),
                    'dishes': []
                }
                
                # 处理商品信息（订单分类字段）
                if '订单分类' in row and pd.notna(row['订单分类']):
                    dishes_str = str(row['订单分类'])
                    if dishes_str and dishes_str != 'nan':
                        # 如果包含逗号，按逗号分割；否则作为单个商品
                        if ',' in dishes_str:
                            order['dishes'] = [{'name': dish.strip(), 'price': 0} for dish in dishes_str.split(',')]
                        else:
                            order['dishes'] = [{'name': dishes_str.strip(), 'price': 0}]
                        print(f"📦 订单{valid_order_id}商品信息: {dishes_str}")
                
                # 如果有前端操作记录，检查是否需要保护前端操作
                if has_frontend_operation and frontend_operation:
                    # 使用订单编号查找现有订单
                    if existing_order_by_number:
                        excel_status = order['status']
                        current_status = existing_order_by_number['status']
                        expected_status = frontend_operation['new_status']
                        
                        # 如果Excel状态与期望的前端状态不同，保持前端状态
                        if excel_status != expected_status:
                            print(f"🛡️  订单{order_number}前端操作保护: Excel={excel_status}, 期望={expected_status}, 保持前端状态")
                            order['status'] = expected_status
                        else:
                            # 状态一致，但不要立即清除前端操作记录，等待一段时间
                            operation_time = frontend_operation['timestamp']
                            time_diff = datetime.now() - operation_time
                            
                            # 如果前端操作时间超过5分钟，才清除记录
                            if time_diff.total_seconds() > 300:  # 5分钟
                                del frontend_operations[str(order_number)]
                                print(f"✅ 订单{order_number}状态同步且操作时间超过5分钟，清除前端操作记录")
                            else:
                                print(f"⏳ 订单{order_number}状态同步，但操作时间较短({time_diff.total_seconds():.0f}秒)，保持保护")
                    else:
                        # 新订单，清除前端操作记录
                        if str(order_number) in frontend_operations:
                            del frontend_operations[str(order_number)]
                        print(f"✅ 新订单{order_number}，清除前端操作记录")
                else:
                    # 没有前端操作记录，检查是否需要保护现有状态
                    if existing_order_by_number:
                        excel_status = order['status']
                        current_status = existing_order_by_number['status']
                        
                        # 如果现有订单状态与Excel状态不同，保持现有状态
                        if excel_status != current_status:
                            print(f"🛡️  订单{order_number}状态保护: Excel={excel_status}, 内存={current_status}, 保持内存状态")
                            order['status'] = current_status
                            
                            # 如果状态被保护，重新记录前端操作
                            if str(order_number) not in frontend_operations:
                                frontend_operations[str(order_number)] = {
                                    'timestamp': datetime.now(),
                                    'old_status': excel_status,
                                    'new_status': current_status,
                                    'protected': True
                                }
                                print(f"📝 重新记录保护操作: 订单{order_number} {excel_status}→{current_status}")
                
                orders_db.append(order)
                valid_order_id += 1
                
            except Exception as e:
                print(f"处理第{index + 1}行数据时出错: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"成功读取 {len(orders_db)} 个订单")
        
    except Exception as e:
        print(f"读取咖啡订单Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()
        # 如果读取失败，保持现有数据不变

def background_excel_reader():
    """后台Excel读取线程"""
    while True:
        try:
            read_excel_orders()
            print(f"Excel数据刷新完成，当前订单数量: {len(orders_db)}")
            
            # 处理完成后，清理锁文件和新数据标志
            try:
                lock_file_path = os.path.join(EXCEL_FOLDER, "data.lock")
                if os.path.exists(lock_file_path):
                    os.remove(lock_file_path)
                    print("🔓 已释放前端处理锁")
                
                flag_file_path = os.path.join(EXCEL_FOLDER, "new_data_ready.flag")
                if os.path.exists(flag_file_path):
                    os.remove(flag_file_path)
                    print("🚩 已移除新数据标志")
            except Exception as e:
                print(f"⚠️ 清理标志文件失败: {e}")
                
        except Exception as e:
            print(f"后台Excel读取出错: {e}")
        
        # 使用配置的刷新间隔
        time.sleep(app.config['DATA_REFRESH_INTERVAL'])

def get_orders_by_status(status=None):
    """根据状态获取订单"""
    if status is None:
        return orders_db
    
    # 处理字符串状态和数字状态的兼容性
    if isinstance(status, str):
        # 字符串状态直接匹配
        return [order for order in orders_db if order['status'] == status]
    else:
        # 数字状态匹配
        return [order for order in orders_db if order['status'] == status]

def update_order_status(order_id, new_status):
    """更新订单状态"""
    global orders_db, frontend_operations, is_excel_updating
    
    print(f"🔍 开始更新订单状态: order_id={order_id} (类型: {type(order_id)}), new_status={new_status}")
    
    # 检查Excel是否正在被外部程序更新
    if is_excel_updating:
        print(f"⚠️  订单{order_id}状态更新被拒绝：咖啡订单Excel文件正在被外部程序更新")
        return False, "系统繁忙，请稍后再试"
    
    # 首先尝试通过订单编号查找（因为前端现在传递的是订单编号）
    print(f"🔍 通过订单编号查找: {order_id}")
    for order in orders_db:
        if str(order['number']) == str(order_id):
            old_status = order['status']
            order['status'] = new_status
            # 记录前端操作时间戳和旧状态（使用订单编号作为键）
            frontend_operations[str(order['number'])] = {
                'timestamp': datetime.now(),
                'old_status': old_status,
                'new_status': new_status
            }
            print(f"📝 记录前端操作: 订单{order['number']} {old_status}→{new_status}")
            # 同步更新Excel文件
            update_excel_order_status_by_number(order['number'], new_status)
            print(f"📝 调用update_excel_order_status_by_number: order_number={order['number']}, new_status={new_status}")
            return True, "操作成功"
    
    # 如果通过订单编号没找到，尝试通过ID查找（向后兼容）
    print(f"🔍 通过订单编号未找到，尝试通过ID查找: {order_id}")
    try:
        order_id_int = int(order_id)
        for order in orders_db:
            if order['id'] == order_id_int:
                old_status = order['status']
                order['status'] = new_status
                # 记录前端操作时间戳和旧状态（使用订单编号作为键）
                frontend_operations[str(order['number'])] = {
                    'timestamp': datetime.now(),
                    'old_status': old_status,
                    'new_status': new_status
                }
                print(f"📝 记录前端操作: 订单{order['number']} {old_status}→{new_status}")
                # 同步更新Excel文件
                update_excel_order_status_by_number(order['number'], new_status)
                return True, "操作成功"
    except ValueError:
        print(f"❌ order_id '{order_id}' 无法转换为整数")
    
    print(f"❌ 未找到订单: {order_id}")
    return False, "订单不存在"

def update_excel_order_status_by_number(order_number, new_status):
    """通过订单编号更新咖啡订单Excel文件中的订单状态"""
    # 使用线程锁防止并发写入冲突
    with excel_write_lock:
        max_retries = app.config['EXCEL_RETRY_COUNT']
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                print(f"🔄 尝试更新Excel文件 (第{retry_count + 1}次): 订单{order_number}")
                
                # 查找最新的咖啡订单Excel文件
                excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
                if not excel_files:
                    print("❌ 未找到咖啡订单Excel文件，无法更新状态")
                    return False
                
                latest_file = max(excel_files, key=os.path.getctime)
                print(f"📁 更新咖啡订单Excel文件: {latest_file}")
                
                # 检查文件是否存在
                if not os.path.exists(latest_file):
                    print(f"❌ Excel文件不存在: {latest_file}")
                    return False
                
                # 检查文件是否可写
                if not os.access(latest_file, os.W_OK):
                    print(f"⚠️  咖啡订单Excel文件无写入权限，请检查文件是否被占用或设置为只读")
                    print(f"   文件路径: {latest_file}")
                    print(f"   建议操作:")
                    print(f"   1. 关闭可能打开该文件的Excel程序")
                    print(f"   2. 右键文件 -> 属性 -> 取消勾选'只读'")
                    print(f"   3. 以管理员身份运行程序")
                    
                    # 尝试修复文件权限
                    try:
                        import stat
                        current_attrs = os.stat(latest_file).st_mode
                        new_attrs = current_attrs | stat.S_IWRITE
                        os.chmod(latest_file, new_attrs)
                        print(f"✅ 已尝试修复文件权限")
                    except Exception as perm_e:
                        print(f"❌ 修复文件权限失败: {perm_e}")
                    
                    if retry_count < max_retries - 1:
                        retry_count += 1
                        time.sleep(app.config['EXCEL_RETRY_DELAY'])  # 使用配置的重试延迟
                        continue
                    else:
                        return False
                
                # 先读取Excel文件，找到对应的订单
                try:
                    df = pd.read_excel(latest_file, engine='openpyxl')
                    print(f"📊 Excel文件包含 {len(df)} 行数据")
                except Exception as read_e:
                    print(f"❌ 读取Excel文件失败: {read_e}")
                    if retry_count < max_retries - 1:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    else:
                        return False
                
                print(f"🎯 目标订单编号: {order_number}")
                
                # 在Excel中查找对应的行
                target_row_index = None
                for index, row in df.iterrows():
                    excel_order_number = row.get('订单编号')
                    if not pd.isna(excel_order_number) and str(excel_order_number) == str(order_number):
                        target_row_index = index
                        break
                
                if target_row_index is None:
                    print(f"❌ 在Excel中未找到订单号 {order_number}")
                    # 打印所有订单编号用于调试
                    print(f"📋 Excel中的所有订单编号:")
                    for index, row in df.iterrows():
                        excel_order_number = row.get('订单编号')
                        if not pd.isna(excel_order_number):
                            print(f"   行{index+2}: {excel_order_number}")
                    return False
                
                excel_row_number = target_row_index + 2  # +2 因为Excel从1开始且有标题行
                print(f"🔍 订单 {order_number} 在Excel第 {excel_row_number} 行")
                
                # 使用openpyxl加载工作簿进行更新
                try:
                    workbook = load_workbook(latest_file)
                    worksheet = workbook.active
                except Exception as load_e:
                    print(f"❌ 加载Excel工作簿失败: {load_e}")
                    if retry_count < max_retries - 1:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    else:
                        return False
                
                # 处理状态更新
                if isinstance(new_status, str):
                    status_text = new_status
                else:
                    status_mapping = {
                        PENDING: '备货中',
                        COMPLETED: '已完成'
                    }
                    status_text = status_mapping.get(new_status, '备货中')
                
                print(f"🔧 更新Excel: 订单{order_number}, 状态{new_status} -> 文本状态'{status_text}'")
                
                # 更新状态列（状态名称是第9列，I列）
                status_cell = worksheet.cell(row=excel_row_number, column=9)
                status_cell.value = status_text
                
                # 同时更新状态代码列（状态是第8列，H列）
                status_code_cell = worksheet.cell(row=excel_row_number, column=8)
                if status_text == '已完成':
                    status_code_cell.value = '5'
                else:
                    status_code_cell.value = '2'
                
                # 保存文件
                print(f"💾 正在保存文件: {latest_file}")
                try:
                    workbook.save(latest_file)
                    print(f"✅ 文件保存成功")
                except Exception as save_e:
                    print(f"❌ 保存文件失败: {save_e}")
                    if retry_count < max_retries - 1:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    else:
                        return False
                
                # 验证文件是否真的被更新了
                print(f"🔍 验证文件更新...")
                try:
                    verification_df = pd.read_excel(latest_file, engine='openpyxl')
                    verification_row = verification_df.iloc[target_row_index]
                    actual_status = verification_row.get('状态名称', 'N/A')
                    actual_status_code = verification_row.get('状态', 'N/A')
                    
                    print(f"📊 验证结果:")
                    print(f"   期望状态: {status_text}")
                    print(f"   实际状态: {actual_status}")
                    print(f"   期望状态码: {'5' if status_text == '已完成' else '2'}")
                    print(f"   实际状态码: {actual_status_code}")
                    
                    if actual_status == status_text:
                        print(f"✅ 文件更新验证成功！")
                        print(f"✅ 成功更新咖啡订单Excel文件，订单{order_number}状态改为{status_text}")
                        return True
                    else:
                        print(f"❌ 文件更新验证失败！状态未正确更新")
                        if retry_count < max_retries - 1:
                            retry_count += 1
                            time.sleep(2)
                            continue
                        else:
                            return False
                except Exception as verify_e:
                    print(f"❌ 验证文件更新失败: {verify_e}")
                    if retry_count < max_retries - 1:
                        retry_count += 1
                        time.sleep(2)
                        continue
                    else:
                        return False
                        
            except PermissionError as e:
                print(f"❌ 咖啡订单Excel文件权限错误: {e}")
                print(f"   请确保Excel文件未被其他程序打开，且具有写入权限")
                if retry_count < max_retries - 1:
                    retry_count += 1
                    time.sleep(2)
                    continue
                else:
                    return False
            except Exception as e:
                print(f"❌ 更新咖啡订单Excel文件时出错: {e}")
                import traceback
                traceback.print_exc()
                if retry_count < max_retries - 1:
                    retry_count += 1
                    time.sleep(2)
                    continue
                else:
                    return False
        
        print(f"❌ 更新Excel文件失败，已重试{max_retries}次")
        return False

def update_excel_order_status(order_id, new_status):
    """更新咖啡订单Excel文件中的订单状态"""
    try:
        # 查找最新的咖啡订单Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        if not excel_files:
            print("❌ 未找到咖啡订单Excel文件，无法更新状态")
            return False
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📁 更新咖啡订单Excel文件: {latest_file}")
        
        # 检查文件是否可写
        if not os.access(latest_file, os.W_OK):
            print(f"⚠️  咖啡订单Excel文件无写入权限，请检查文件是否被占用或设置为只读")
            print(f"   文件路径: {latest_file}")
            print(f"   建议操作:")
            print(f"   1. 关闭可能打开该文件的Excel程序")
            print(f"   2. 右键文件 -> 属性 -> 取消勾选'只读'")
            print(f"   3. 以管理员身份运行程序")
            return False
        
        # 先读取Excel文件，找到对应的订单
        df = pd.read_excel(latest_file, engine='openpyxl')
        print(f"📊 Excel文件包含 {len(df)} 行数据")
        
        # 找到对应的订单（通过order_id在orders_db中的索引）
        if order_id <= 0 or order_id > len(orders_db):
            print(f"❌ 订单ID {order_id} 超出orders_db范围 (1-{len(orders_db)})")
            return False
        
        target_order = orders_db[order_id - 1]  # order_id从1开始，转换为0基索引
        target_order_number = target_order['number']
        print(f"🎯 目标订单: {target_order_number} (ID: {order_id})")
        
        # 验证目标订单是否正确
        print(f"🔍 验证目标订单: 期望更新订单编号 {target_order_number}")
        
        # 打印orders_db中的所有订单，用于调试
        print(f"📋 orders_db中的订单:")
        for i, order in enumerate(orders_db):
            print(f"   ID {i+1}: {order['number']} - {order['status']}")
        
        # 在Excel中查找对应的行
        target_row_index = None
        for index, row in df.iterrows():
            excel_order_number = row.get('订单编号')
            if not pd.isna(excel_order_number) and str(excel_order_number) == str(target_order_number):
                target_row_index = index
                break
        
        if target_row_index is None:
            print(f"❌ 在Excel中未找到订单号 {target_order_number}")
            return False
        
        excel_row_number = target_row_index + 2  # +2 因为Excel从1开始且有标题行
        print(f"🔍 订单 {target_order_number} 在Excel第 {excel_row_number} 行")
        
        # 打印Excel中的所有订单，用于调试
        print(f"📋 Excel中的订单:")
        for index, row in df.iterrows():
            excel_order_number = row.get('订单编号')
            if not pd.isna(excel_order_number):
                print(f"   行{index+2}: {excel_order_number}")
        
        # 使用openpyxl加载工作簿进行更新
        workbook = load_workbook(latest_file)
        worksheet = workbook.active
        
        # 处理状态更新
        if isinstance(new_status, str):
            status_text = new_status
        else:
            status_mapping = {
                PENDING: '备货中',
                COMPLETED: '已完成'
            }
            status_text = status_mapping.get(new_status, '备货中')
        
        print(f"🔧 更新Excel: 订单{target_order_number}, 状态{new_status} -> 文本状态'{status_text}'")
        
        # 更新状态列（状态名称是第9列，I列）
        status_cell = worksheet.cell(row=excel_row_number, column=9)
        status_cell.value = status_text
        
        # 同时更新状态代码列（状态是第8列，H列）
        status_code_cell = worksheet.cell(row=excel_row_number, column=8)
        if status_text == '已完成':
            status_code_cell.value = '5'
        else:
            status_code_cell.value = '2'
        
        # 保存文件
        print(f"💾 正在保存文件: {latest_file}")
        workbook.save(latest_file)
        
        # 验证文件是否真的被更新了
        print(f"🔍 验证文件更新...")
        verification_df = pd.read_excel(latest_file, engine='openpyxl')
        verification_row = verification_df.iloc[target_row_index]
        actual_status = verification_row.get('状态名称', 'N/A')
        actual_status_code = verification_row.get('状态', 'N/A')
        
        print(f"📊 验证结果:")
        print(f"   期望状态: {status_text}")
        print(f"   实际状态: {actual_status}")
        print(f"   期望状态码: {'5' if status_text == '已完成' else '2'}")
        print(f"   实际状态码: {actual_status_code}")
        
        if actual_status == status_text:
            print(f"✅ 文件更新验证成功！")
        else:
            print(f"❌ 文件更新验证失败！状态未正确更新")
        
        print(f"✅ 成功更新咖啡订单Excel文件，订单{target_order_number}状态改为{status_text}")
        return True
            
    except PermissionError as e:
        print(f"❌ 咖啡订单Excel文件权限错误: {e}")
        print(f"   请确保Excel文件未被其他程序打开，且具有写入权限")
        return False
    except Exception as e:
        print(f"❌ 更新咖啡订单Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

# 错误处理中间件
@app.errorhandler(500)
def internal_error(error):
    """处理内部服务器错误"""
    print(f"❌ 服务器内部错误: {error}")
    return jsonify({'code': 0, 'msg': '服务器内部错误，请稍后重试'}), 500

@app.errorhandler(503)
def service_unavailable(error):
    """处理服务不可用错误"""
    print(f"❌ 服务不可用: {error}")
    return jsonify({'code': 0, 'msg': '服务暂时不可用，请稍后重试'}), 503

# Flask路由定义
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/orders')
def api_orders():
    """获取未完成订单"""
    # 获取所有非"已完成"状态的订单
    pending_orders = [order for order in orders_db if order['status'] != '已完成']
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': pending_orders
    })

@app.route('/api/orders/all')
def api_all_orders():
    """获取所有订单"""
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': orders_db
    })

@app.route('/api/order/<order_id>/<action>', methods=['POST'])
def api_update_order(order_id, action):
    """更新订单状态"""
    print(f"🔔 收到API请求: 订单{order_id}, 动作: {action}")
    
    # 添加请求频率限制检查
    current_time = time.time()
    request_key = f"{order_id}_{action}"
    
    # 检查是否在配置的时间内重复请求
    if hasattr(app, 'last_requests'):
        if request_key in app.last_requests:
            last_time = app.last_requests[request_key]
            if current_time - last_time < app.config['REQUEST_RATE_LIMIT']:
                print(f"⚠️  请求过于频繁: 订单{order_id}, 动作: {action}")
                return jsonify({'code': 0, 'msg': '请求过于频繁，请稍后再试'})
    else:
        app.last_requests = {}
    
    app.last_requests[request_key] = current_time
    
    try:
        if action == 'complete':
            print(f"🔄 开始处理完成订单请求: 订单{order_id}")
            
            # 检查Excel是否正在被外部程序更新
            if is_excel_updating:
                print(f"⚠️  订单{order_id}状态更新被拒绝：咖啡订单Excel文件正在被外部程序更新")
                return jsonify({'code': 0, 'msg': '系统繁忙，请稍后再试'})
            
            success, message = update_order_status(order_id, '已完成')
            print(f"📊 处理结果: 成功={success}, 消息={message}")
            
            if success:
                response_data = {
                    'code': 1, 
                    'msg': f'{message}，咖啡订单Excel文件已同步更新'
                }
                print(f"✅ 返回成功响应: {response_data}")
                return jsonify(response_data)
            else:
                response_data = {
                    'code': 0, 
                    'msg': message
                }
                print(f"❌ 返回失败响应: {response_data}")
                return jsonify(response_data)

        else:
            print(f"❌ 无效的操作: {action}")
            return jsonify({'code': 0, 'msg': '无效的操作'})
            
    except Exception as e:
        print(f"💥 API处理异常: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'code': 0, 'msg': f'操作失败: {str(e)}'})

@app.route('/api/statistics')
def api_statistics():
    """获取统计信息"""
    total_orders = len(orders_db)
    # 统计未完成订单（非"已完成"状态的订单）
    pending_orders = len([order for order in orders_db if order['status'] != '已完成'])
    # 统计已完成订单
    completed_orders = len([order for order in orders_db if order['status'] == '已完成'])
    total_amount = sum(order['amount'] for order in orders_db)
    
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': {
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'completed_orders': completed_orders,
            'total_amount': round(total_amount, 2)
        }
    })

@app.route('/api/orders/<status>')
def api_orders_by_status(status):
    """根据状态获取订单"""
    try:
        status_code = int(status)
        filtered_orders = get_orders_by_status(status_code)
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': filtered_orders
        })
    except ValueError:
        return jsonify({'code': 0, 'msg': '无效的状态码'})

@app.route('/api/excel-info')
def api_excel_info():
    """获取咖啡订单Excel文件信息"""
    try:
        if not ensure_orders_folder():
            return jsonify({
                'code': 0,
                'msg': '无法访问桌面路径',
                'data': None
            })
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if excel_files:
            latest_file = max(excel_files, key=os.path.getctime)
            file_info = {
                'file_name': os.path.basename(latest_file),
                'file_path': latest_file,
                'last_modified': datetime.fromtimestamp(os.path.getmtime(latest_file)).isoformat(),
                'file_size': os.path.getsize(latest_file),
                'order_count': len(orders_db)
            }
        else:
            file_info = {
                'file_name': '无文件',
                'file_path': '',
                'last_modified': '',
                'file_size': 0,
                'order_count': 0
            }
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': file_info
        })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取文件信息失败: {str(e)}'})

@app.route('/api/excel-status')
def api_excel_status():
    """获取咖啡订单Excel文件状态信息"""
    try:
        if not ensure_orders_folder():
            return jsonify({
                'code': 0,
                'msg': '无法访问桌面路径',
                'data': None
            })
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if excel_files:
            latest_file = max(excel_files, key=os.path.getctime)
            # 读取咖啡订单Excel文件获取最新状态
            df = pd.read_excel(latest_file, engine='openpyxl')
            
            # 统计各状态数量
            status_counts = df['订单状态'].value_counts().to_dict()
            
            return jsonify({
                'code': 1,
                'msg': 'success',
                'data': {
                    'file_name': os.path.basename(latest_file),
                    'last_modified': datetime.fromtimestamp(os.path.getmtime(latest_file)).isoformat(),
                    'status_counts': status_counts,
                    'total_orders': len(df)
                }
            })
        else:
            return jsonify({
                'code': 0,
                'msg': '未找到咖啡订单Excel文件',
                'data': None
            })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取咖啡订单Excel状态失败: {str(e)}'})

@app.route('/api/frontend-operations')
def api_frontend_operations():
    """获取前端操作记录"""
    try:
        operations_info = {}
        for order_id, operation_data in frontend_operations.items():
            if isinstance(operation_data, dict):
                operations_info[order_id] = {
                    'operation_time': operation_data['timestamp'].isoformat(),
                    'time_ago': str(datetime.now() - operation_data['timestamp']),
                    'old_status': operation_data['old_status'],
                    'new_status': operation_data['new_status']
                }
            else:
                # 兼容旧格式
                operations_info[order_id] = {
                    'operation_time': operation_data.isoformat(),
                    'time_ago': str(datetime.now() - operation_data)
                }
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': {
                'operations_count': len(frontend_operations),
                'operations': operations_info
            }
        })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取前端操作记录失败: {str(e)}'})

@app.route('/api/search/pickup-code/<pickup_code>')
def api_search_by_pickup_code(pickup_code):
    """根据取餐码查询订单"""
    try:
        # 在所有订单中查找匹配的取餐码
        matching_orders = []
        
        for order in orders_db:
            # 从remark字段中提取取餐码
            remark = order.get('remark', '')
            if '取餐码:' in remark:
                order_pickup_code = remark.split('取餐码:')[1].strip()
                if order_pickup_code == pickup_code:
                    matching_orders.append(order)
        
        if matching_orders:
            return jsonify({
                'code': 1,
                'msg': f'找到 {len(matching_orders)} 个匹配的订单',
                'data': matching_orders
            })
        else:
            return jsonify({
                'code': 0,
                'msg': f'未找到取餐码为 {pickup_code} 的订单',
                'data': []
            })
            
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'查询失败: {str(e)}'})

@app.route('/api/search/phone/<phone>')
def api_search_by_phone(phone):
    """根据手机号查询订单"""
    try:
        # 在所有订单中查找匹配的手机号
        matching_orders = []
        
        for order in orders_db:
            if order.get('phone', '') == phone:
                matching_orders.append(order)
        
        if matching_orders:
            return jsonify({
                'code': 1,
                'msg': f'找到 {len(matching_orders)} 个匹配的订单',
                'data': matching_orders
            })
        else:
            return jsonify({
                'code': 0,
                'msg': f'未找到手机号为 {phone} 的订单',
                'data': []
            })
            
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'查询失败: {str(e)}'})

@app.route('/api/system-status')
def api_system_status():
    """获取系统状态"""
    try:
        # 安全处理时间戳
        excel_time_str = None
        if excel_file_modified_time is not None:
            try:
                if hasattr(excel_file_modified_time, 'isoformat'):
                    excel_time_str = excel_file_modified_time.isoformat()
                else:
                    # 如果是float类型的时间戳，转换为datetime
                    from datetime import datetime
                    excel_time_str = datetime.fromtimestamp(excel_file_modified_time).isoformat()
            except Exception:
                excel_time_str = str(excel_file_modified_time)
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': {
                'is_excel_updating': is_excel_updating,
                'excel_file_modified_time': excel_time_str,
                'frontend_operations_count': len(frontend_operations),
                'orders_count': len(orders_db)
            }
        })
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取系统状态失败: {str(e)}'})

# 导出函数供main.py使用
def init_app():
    """初始化应用"""
    print("📊 初始化数据读取...")
    
    # 确保Excel文件可写
    ensure_excel_files_writable()
    
    read_excel_orders()
    
    # 启动后台Excel读取线程
    excel_thread = threading.Thread(target=background_excel_reader, daemon=True)
    excel_thread.start()
    print("🔄 后台Excel读取线程已启动，每分钟刷新一次")

def run_app():
    """运行Flask应用"""
    # 使用配置文件中的参数
    app.run(
        debug=app.config['DEBUG'],
        host=app.config['HOST'],
        port=app.config['PORT'],
        threaded=app.config['THREADED'],
        processes=app.config['PROCESSES']
    ) 